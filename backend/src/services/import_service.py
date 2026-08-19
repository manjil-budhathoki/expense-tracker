import os
import csv
import io
import shutil
import datetime
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from src.services.services import _resolve_payment_method
from src.models.model import ExpenseModel, CategoryModel, ImportFileModel
from src.schemas.schema import TransactionType, PaymentMethod

UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


def _get_or_create_category(db: Session, name: str) -> int:
    name = name.strip()
    category = db.query(CategoryModel).filter(CategoryModel.name == name).first()
    if category:
        return category.id
    category = CategoryModel(name=name)
    db.add(category)
    db.commit()
    db.refresh(category)
    return category.id


def _parse_row(db: Session, row: dict):
    return ExpenseModel(
        date=datetime.date.fromisoformat(str(row["Date"])),
        category_id=_get_or_create_category(db, row["Category"]),
        type=TransactionType(str(row["Type"]).strip().lower()),
        payment_method=_resolve_payment_method(row["Payment Method"]),
        amount=float(row["Amount"]),
        note=row.get("Note") or None,
    )


def _parse_csv_rows(file_bytes: bytes):
    text = file_bytes.decode("utf-8")
    return list(csv.DictReader(io.StringIO(text)))


def _parse_xlsx_rows(file_bytes: bytes):
    wb = load_workbook(io.BytesIO(file_bytes))
    ws = wb.active
    header = [cell.value for cell in ws[1]]
    rows = []
    for row_cells in ws.iter_rows(min_row=2):
        row = {header[j]: cell.value for j, cell in enumerate(row_cells)}
        if row.get("Date") and hasattr(row["Date"], "isoformat"):
            row["Date"] = row["Date"].isoformat()
        rows.append(row)
    return rows


def create_import(db: Session, name: str, file, file_format: str):
    existing = db.query(ImportFileModel).filter(ImportFileModel.name == name).first()
    if existing:
        raise ValueError(f"An import named '{name}' already exists.")

    # Read once for parsing
    file_bytes = file.file.read()

    if file_format == "csv":
        raw_rows = _parse_csv_rows(file_bytes)
    elif file_format == "xlsx":
        raw_rows = _parse_xlsx_rows(file_bytes)
    else:
        raise ValueError("Unsupported format")

    created, errors = [], []
    for i, row in enumerate(raw_rows, start=2):  # row 1 is header
        try:
            expense = _parse_row(db, row)
            db.add(expense)
            created.append(expense)
        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    db.commit()

    # Save physical file
    file_path = os.path.join(UPLOAD_DIR, f"{name}.{file_format}")
    with open(file_path, "wb") as buffer:
        buffer.write(file_bytes)

    # Save metadata record
    record = ImportFileModel(
        name=name,
        file_path=file_path,
        format=file_format,
        rows_created=len(created),
    )
    db.add(record)
    db.commit()
    db.refresh(record)

    return record, len(created), errors